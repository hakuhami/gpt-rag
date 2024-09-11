import pandas as pd
import os
import json
from PIL import Image
import pdf2image
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# popplerのパスを指定（自分の環境に合わせて変更してください）
POPPLER_PATH = r"C:\Users\hakusen_shu\OneDrive\データセット\確定版データセット\Experiment\poppler-24.07.0\Library\bin"

def load_url_pdf_mapping(json_path):
    with open(json_path, 'r', encoding='utf-8-sig') as f:
        mapping = json.load(f)
    return {v: k for k, v in mapping.items()}  # PDFファイル名をキー、URLを値とする辞書に変換

def pdf_to_image(pdf_path, page_number, dpi=200):
    """PDFの特定のページを画像に変換する"""
    images = pdf2image.convert_from_path(pdf_path, first_page=page_number, last_page=page_number, dpi=dpi, poppler_path=POPPLER_PATH)
    return images[0]

def add_image_to_excel(ws, img, row, col):
    """ExcelワークシートにPIL画像を追加する"""
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr = img_byte_arr.getvalue()

    img = XLImage(io.BytesIO(img_byte_arr))
    ws.add_image(img, f'{col}{row}')

def process_tsv_and_add_to_excel(wb, tsv_path, pdf_dir, sheet_name, url_pdf_mapping):
    # TSVファイルを読み込む
    df = pd.read_csv(tsv_path, sep='\t')

    # "N/A"をNoneに置換しないようにする
    df = df.fillna("N/A")

    # URLカラムを追加
    df['URL'] = df['PDF'].map(url_pdf_mapping)

    # 新しいワークシートを作成
    ws = wb.create_sheet(sheet_name)

    # ヘッダーを追加
    headers = list(df.columns) + ['Image']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 各行を処理
    for index, row in df.iterrows():
        excel_row = index + 2  # Excelは1-indexedで、1行目はヘッダー

        # TSVデータをExcelに追加
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # PDFから画像を抽出
        pdf_path = os.path.join(pdf_dir, row['PDF'])
        page_number = int(row['page_number'])
        img = pdf_to_image(pdf_path, page_number, dpi=200)

        # 画像をリサイズ（アスペクト比を維持）
        max_width = 600  # 最大幅を設定
        width_percent = (max_width / float(img.size[0]))
        new_height = int((float(img.size[1]) * float(width_percent)))
        img = img.resize((max_width, new_height), Image.LANCZOS)
        print(f"Extracted image from {pdf_path} page {page_number}")

        # 画像をExcelに追加
        img_col = get_column_letter(len(row) + 1)
        add_image_to_excel(ws, img, excel_row, img_col)

        # 行の高さを調整
        ws.row_dimensions[excel_row].height = max(new_height * 0.75, 100)  # ピクセルからポイントに変換、最小高さを設定

    # 列幅を調整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max((max_length + 2) * 1.2, 15)  # 最小幅を設定
        ws.column_dimensions[column_letter].width = adjusted_width

    # 画像列の幅を設定
    ws.column_dimensions[img_col].width = 80  # 適切な幅に調整

def process_directory_and_create_excel(input_dir, pdf_dir, output_excel_path, url_pdf_mapping):
    # Excelワークブックを作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除

    # ディレクトリ内の各TSVファイルを処理
    for tsv_file in os.listdir(input_dir):
        if tsv_file.endswith('.tsv'):
            tsv_path = os.path.join(input_dir, tsv_file)
            sheet_name = os.path.splitext(tsv_file)[0]  # 拡張子を除いたファイル名
            process_tsv_and_add_to_excel(wb, tsv_path, pdf_dir, sheet_name, url_pdf_mapping)

    # Excelファイルを保存
    wb.save(output_excel_path)

# スクリプトを実行
if __name__ == "__main__":
    input_dir = "./data/processed/statistics/non-coresspond/Korean_withRAG"
    pdf_dir = "./data/raw/PDFs"
    output_excel_path = "./data/processed/statistics/non-coresspond/Korean_withRAG_analysis.xlsx"
    url_pdf_json_path = "./data/raw/PDFs/.Korean_URL-PDF_List.json"

    # URL-PDF対応マッピングを読み込む
    url_pdf_mapping = load_url_pdf_mapping(url_pdf_json_path)

    process_directory_and_create_excel(input_dir, pdf_dir, output_excel_path, url_pdf_mapping)
    print(f"Analysis completed. Excel file saved at: {output_excel_path}")