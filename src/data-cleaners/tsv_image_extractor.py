import pandas as pd
import os
from PIL import Image
import pdf2image
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

POPPLER_PATH = r"C:\Users\hakusen_shu\OneDrive\データセット\確定版データセット\Experiment\poppler-24.07.0\Library\bin"

def pdf_to_image(pdf_path, page_number):
    """PDFの特定のページを画像に変換する"""
    images = pdf2image.convert_from_path(pdf_path, first_page=page_number, last_page=page_number, poppler_path=POPPLER_PATH)
    return images[0]

def add_image_to_excel(ws, img, row, col):
    """ExcelワークシートにPIL画像を追加する"""
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr = img_byte_arr.getvalue()

    img = XLImage(io.BytesIO(img_byte_arr))
    ws.add_image(img, f'{col}{row}')

def process_tsv_and_add_to_excel(wb, tsv_path, pdf_dir, sheet_name):
    # TSVファイルを読み込む
    df = pd.read_csv(tsv_path, sep='\t')

    # 新しいワークシートを作成
    ws = wb.create_sheet(sheet_name)

    # ヘッダーを追加
    headers = list(df.columns) + ['Image']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # 各行を処理
    for index, row in df.iterrows():
        excel_row = index + 2  # Excelは1-indexedで、1行目はヘッダー

        # TSVデータをExcelに追加
        for col, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=col, value=value)

        # PDFから画像を抽出
        pdf_path = os.path.join(pdf_dir, row['PDF'])
        page_number = int(row['page_number'])
        img = pdf_to_image(pdf_path, page_number)

        # 画像をリサイズ（必要に応じて）
        img.thumbnail((300, 300))  # 最大サイズを300x300にする
        print(f"Extracted image from {pdf_path} page {page_number}")

        # 画像をExcelに追加
        add_image_to_excel(ws, img, excel_row, chr(ord('A') + len(row)))

        # 行の高さを調整
        ws.row_dimensions[excel_row].height = 150  # ピクセル単位
        print(f"Added image to row {excel_row}")

    # 列幅を調整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 画像列の幅を設定
    ws.column_dimensions[chr(ord('A') + len(df.columns))].width = 40

def process_directory_and_create_excel(input_dir, pdf_dir, output_excel_path):
    # Excelワークブックを作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除

    # ディレクトリ内の各TSVファイルを処理
    for tsv_file in os.listdir(input_dir):
        if tsv_file.endswith('.tsv'):
            tsv_path = os.path.join(input_dir, tsv_file)
            sheet_name = os.path.splitext(tsv_file)[0]  # 拡張子を除いたファイル名
            process_tsv_and_add_to_excel(wb, tsv_path, pdf_dir, sheet_name)

    # Excelファイルを保存
    wb.save(output_excel_path)

# スクリプトを実行
if __name__ == "__main__":
    input_dir = "./data/processed/statistics/non-coresspond/Chinese_withRAG"
    pdf_dir = "./data/raw/PDFs"
    output_excel_path = "./data/processed/statistics/non-coresspond/Chinese_withRAG_analysis.xlsx"

    process_directory_and_create_excel(input_dir, pdf_dir, output_excel_path)
    print(f"Analysis completed. Excel file saved at: {output_excel_path}")